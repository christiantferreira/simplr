"""
XLSX Export for working papers, roll-forward reports, and generic JE export.
Uses openpyxl (MIT License).
"""

import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from ..calc.decimal_utils import round_penny, ZERO


# Style constants
HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E1052", end_color="1E1052", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Inter", size=10)
MONEY_FONT = Font(name="Inter", size=10)
MONEY_FORMAT = '#,##0.00'
DATE_FORMAT = 'YYYY-MM-DD'

TOTAL_FONT = Font(name="Inter", bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="F0EDFB", end_color="F0EDFB", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)


def export_roll_forward_xlsx(
    report_data: dict,
    client_name: str,
    fiscal_year_end: str,
    report_date: date,
    period_start: date,
    period_end: date,
) -> bytes:
    """
    Export a roll-forward report as XLSX.

    report_data format:
    {
        "contracts": [
            {
                "description": str,
                "account": str,
                "opening_balance": Decimal,
                "additions": Decimal,
                "amortization": Decimal,
                "ending_balance": Decimal,
                "next_je_amount": Decimal,
                "next_je_month": str,
            },
            ...
        ],
        "totals": {
            "opening_balance": Decimal,
            "additions": Decimal,
            "amortization": Decimal,
            "ending_balance": Decimal,
        }
    }
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roll-Forward Report"

    # Metadata section
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Simplr — Prepaid Expenses Roll-Forward"
    ws["A1"].font = Font(name="Inter", bold=True, size=14, color="7C6AED")

    ws["A2"] = "Client:"
    ws["B2"] = client_name
    ws["A3"] = "Fiscal Year-End:"
    ws["B3"] = fiscal_year_end
    ws["A4"] = "Report Period:"
    ws["B4"] = f"{period_start.strftime('%b %d, %Y')} to {period_end.strftime('%b %d, %Y')}"
    ws["A5"] = "Generated:"
    ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for r in range(2, 6):
        ws[f"A{r}"].font = Font(name="Inter", bold=True, size=10)
        ws[f"B{r}"].font = Font(name="Inter", size=10)

    # Headers
    header_row = 7
    headers = [
        "Contract", "Account", "Opening Balance",
        "Additions", "Amortization", "Ending Balance", "Next JE",
    ]
    col_widths = [30, 12, 18, 15, 15, 18, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    row = header_row + 1
    for contract in report_data["contracts"]:
        ws.cell(row=row, column=1, value=contract["description"]).font = BODY_FONT
        ws.cell(row=row, column=2, value=contract["account"]).font = BODY_FONT

        for col_idx, key in enumerate(
            ["opening_balance", "additions", "amortization", "ending_balance"], 3
        ):
            cell = ws.cell(row=row, column=col_idx, value=float(contract[key]))
            cell.font = MONEY_FONT
            cell.number_format = MONEY_FORMAT
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER

        next_je = contract.get("next_je_amount", ZERO)
        next_month = contract.get("next_je_month", "")
        next_str = f"${next_je:,.2f} - {next_month}" if next_je > ZERO else "Complete"
        ws.cell(row=row, column=7, value=next_str).font = BODY_FONT

        for col_idx in range(1, 8):
            ws.cell(row=row, column=col_idx).border = THIN_BORDER

        row += 1

    # Totals row
    totals = report_data["totals"]
    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL

    for col_idx, key in enumerate(
        ["opening_balance", "additions", "amortization", "ending_balance"], 3
    ):
        cell = ws.cell(row=row, column=col_idx, value=float(totals[key]))
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = MONEY_FORMAT
        cell.alignment = Alignment(horizontal="right")
        cell.border = THIN_BORDER

    for col_idx in range(1, 8):
        ws.cell(row=row, column=col_idx).fill = TOTAL_FILL
        ws.cell(row=row, column=col_idx).border = THIN_BORDER

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
